# -*- utf-8 -*-
import openpyxl as xl
import json

class Xlsx2Json:
    def __init__(self, file_path):
        self.wb = xl.load_workbook(file_path)
        self.group_info = self.wb['group_info']
        self.tab_info = self.wb['tab_info']
        self.site_info = self.wb['site_info']
        self.data = []

    def __del__(self):
        self.wb.close()
        pass

    def parse(self):
        kv = {}
        for row in self.group_info.iter_rows(min_row=2, max_row=self.group_info.max_row, min_col=1, max_col=2):
            group_id = row[0].value
            group_title = row[1].value
            kv['group_'+str(group_id)] = len(self.data)
            self.data.append({'id': group_id, 'title': group_title, 'tabs': []})
        for row in self.tab_info.iter_rows(min_row=2, max_row=self.tab_info.max_row, min_col=1, max_col=3):
            tab_id = row[0].value
            group_id = row[1].value
            tab_title = row[2].value
            kv['tab_'+str(tab_id)] = len(self.data[kv['group_'+str(group_id)]]['tabs'])
            self.data[kv['group_'+str(group_id)]]['tabs'].append({'id': tab_id, 'title': tab_title, 'sites': []})
        for row in self.site_info.iter_rows(min_row=2, max_row=self.site_info.max_row, min_col=1, max_col=9):
            site_id = row[0].value
            group_id = row[1].value
            tab_id = row[2].value
            site_title = row[3].value
            description = row[4].value
            url = row[5].value
            icon = row[6].value
            show = bool(row[7].value)
            url_ad = row[8].value
            self.data[kv['group_'+str(group_id)]]['tabs'][kv['tab_'+str(tab_id)]]['sites'].append({'id': site_id, 'title': site_title, 'description': description, 'url': url, 'icon': icon, 'show': show, 'url_ad': url_ad})
        pass
    

    def write2file(self, file_path):
        with open(file_path, 'w') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=4)

if __name__ == '__main__':
    xlsx2json = Xlsx2Json('./datastore.xlsx')
    xlsx2json.parse()
    xlsx2json.write2file('../config/nav.json')
