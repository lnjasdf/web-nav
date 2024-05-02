# -*- utf-8 -*-
import openpyxl as xl
import sqlite3

class Xlsx2Sql:
    def __init__(self, file_path, db_path):
        self.db = sqlite3.connect(db_path)
        self.wb = xl.load_workbook(file_path)
        self.group_info = self.wb['group_info']
        self.tab_info = self.wb['tab_info']
        self.site_info = self.wb['site_info']
        self.data = []
        self.cur = self.db.cursor()

    def save_group(self):
        for row in self.group_info.iter_rows(min_row=2, max_row=self.group_info.max_row, min_col=1, max_col=2):
            sql = 'insert into group_info(title) values (?)'
            self.cur.execute(sql, (row[1].value,))
        pass

    def save_tab(self):
        for row in self.tab_info.iter_rows(min_row=2, max_row=self.tab_info.max_row, min_col=1, max_col=3):
            tab_id = row[0].value
            group_id = row[1].value
            tab_title = row[2].value
            sql = 'insert into tab_info(title, group_id) values (?, ?)'
            self.cur.execute(sql, (tab_title, group_id))

    def save_site(self):
        for row in self.site_info.iter_rows(min_row=2, max_row=self.site_info.max_row, min_col=1, max_col=9):
            site_id = row[0].value
            group_id = row[1].value
            tab_id = row[2].value
            site_title = row[3].value
            description = row[4].value
            url = row[5].value
            icon = row[6].value
            show = bool(row[7].value)
            url_ad = row[8].value
            sql = 'insert into site_info(title, description, url, icon, show, url_ad, group_id, tab_id) values (?, ?, ?, ?, ?, ?, ?, ?)'
            self.cur.execute(sql, (site_title, description, url, icon, show, url_ad, group_id, tab_id))

    def commit(self):
        self.db.commit()

    def close(self):
        self.cur.close()
        self.db.close()


if __name__ == '__main__':
    xlsx2Sql = Xlsx2Sql('./datastore.xlsx', '/Users/liunanji/Documents/project/database/web-nav/data.sqlite')
    xlsx2Sql.save_site()
    xlsx2Sql.commit()
    xlsx2Sql.close()
    
    